import pandas as pd
import numpy as np
import scipy.stats as stats
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def calculate_confidence_interval(successes, total, confidence=0.95):
    """
    Berechnet das Konfidenzintervall für eine Proportion mit der Wilson-Score-Methode.
    """
    if total == 0:
        return 0, 0, 0

    proportion = successes / total
    z = stats.norm.ppf((1 + confidence) / 2)

    # Wilson Score Interval
    denominator = 1 + z ** 2 / total
    center = (proportion + z ** 2 / (2 * total)) / denominator
    margin = z * np.sqrt(proportion * (1 - proportion) / total + z ** 2 / (4 * total ** 2)) / denominator

    lower_bound = max(0, center - margin)
    upper_bound = min(1, center + margin)

    return proportion, lower_bound, upper_bound


def run_ab_test(data_a, data_b, metric_name, min_effect_size=0.05, confidence=0.95):
    """
    Führt einen A/B-Test durch und gibt die Ergebnisse zurück.
    data_a: Dictionary mit 'success' und 'total' für Variante A
    data_b: Dictionary mit 'success' und 'total' für Variante B
    metric_name: Name der Metrik (z.B. 'CTR' oder 'Bounce Rate')
    min_effect_size: Minimale Effektgröße, die als signifikant betrachtet wird
    confidence: Konfidenzniveau (Standard: 0.95)
    """
    # Berechne Konversionsraten und Konfidenzintervalle
    rate_a, ci_lower_a, ci_upper_a = calculate_confidence_interval(data_a['success'], data_a['total'], confidence)
    rate_b, ci_lower_b, ci_upper_b = calculate_confidence_interval(data_b['success'], data_b['total'], confidence)

    # Berechne relative Änderung (für CTR positiv, für Bounce Rate negativ)
    if metric_name == "CTR":
        relative_change = (rate_b - rate_a) / rate_a
        significant_improvement = rate_b > rate_a * (1 + min_effect_size) and ci_lower_b > ci_upper_a
    elif metric_name == "Bounce Rate":
        relative_change = (rate_a - rate_b) / rate_a  # Für Bounce Rate ist eine Reduktion besser
        significant_improvement = rate_b < rate_a * (1 - min_effect_size) and ci_upper_b < ci_lower_a

    # Standardabweichung für beide Varianten berechnen
    # Für Proportionen: Standardabweichung = sqrt(p * (1-p))
    std_dev_a = np.sqrt(rate_a * (1 - rate_a))
    std_dev_b = np.sqrt(rate_b * (1 - rate_b))

    # Z-Score für den Unterschied berechnen
    pooled_std = np.sqrt(std_dev_a ** 2 / data_a['total'] + std_dev_b ** 2 / data_b['total'])
    z_score = (rate_b - rate_a) / pooled_std if pooled_std > 0 else 0

    # p-Wert berechnen
    p_value = 2 * (1 - stats.norm.cdf(abs(z_score)))

    # Erstelle Ergebnisdictionary
    result = {
        'metric': metric_name,
        'A_name': "Applications Open Now",
        'B_name': "Applications Closing Soon",
        'A_rate': rate_a,
        'A_ci_lower': ci_lower_a,
        'A_ci_upper': ci_upper_a,
        'A_sample_size': data_a['total'],
        'B_rate': rate_b,
        'B_ci_lower': ci_lower_b,
        'B_ci_upper': ci_upper_b,
        'B_sample_size': data_b['total'],
        'relative_change': relative_change,
        'min_effect_size': min_effect_size,
        'confidence': confidence,
        'z_score': z_score,
        'p_value': p_value,
        'significant_improvement': significant_improvement
    }

    return result


def create_normal_distribution_data(mean, std_dev, num_points=500, cutoff_threshold=0.01):
    """
    Erstellt Datenpunkte für eine Normalverteilung mit sehr feinen Abstufungen.
    Schneidet die Verteilung bei einem bestimmten Schwellenwert ab, um nur relevante Bereiche einzuschließen.

    Args:
        mean: Mittelwert der Normalverteilung
        std_dev: Standardabweichung der Normalverteilung
        num_points: Anzahl der zu erzeugenden Datenpunkte (höher = feinere Kurve)
        cutoff_threshold: Schwellenwert für y-Werte (alle Werte unter diesem Schwellenwert werden entfernt)

    Returns:
        Tuple aus zwei Listen (x-Werte, y-Werte) für die relevanten Bereiche der Kurve
    """
    # Erstelle zunächst einen breiten Bereich
    range_factor = 4
    x_min = mean - range_factor * std_dev
    x_max = mean + range_factor * std_dev

    # Erzeuge feine Datenpunkte
    x_values = np.linspace(x_min, x_max, num_points)
    y_values = stats.norm.pdf(x_values, mean, std_dev)

    # Finde den maximalen y-Wert
    y_max = np.max(y_values)

    # Berechne den absoluten Schwellenwert basierend auf dem relativen Schwellenwert
    abs_threshold = y_max * cutoff_threshold

    # Filtere x- und y-Werte, wo y über dem Schwellenwert liegt
    mask = y_values > abs_threshold
    filtered_x = x_values[mask]
    filtered_y = y_values[mask]

    return filtered_x, filtered_y


def create_excel_for_think_cell(results, filename="ab_test_results.xlsx"):
    """
    Erstellt eine Excel-Datei mit den Ergebnissen, die für Think-Cell optimiert ist.
    Enthält zusätzlich Daten für Normalverteilungskurven mit sehr feinen Abstufungen.
    """
    wb = Workbook()

    # Erstelle ein Blatt für jede Metrik
    for result in results:
        metric = result['metric']
        ws = wb.create_sheet(title=metric)

        # Erstelle Daten für die Kurven
        data = [
            ['Version', 'Rate', 'Lower CI', 'Upper CI'],
            [result['A_name'], result['A_rate'], result['A_ci_lower'], result['A_ci_upper']],
            [result['B_name'], result['B_rate'], result['B_ci_lower'], result['B_ci_upper']]
        ]

        # Füge Daten zum Arbeitsblatt hinzu
        for row in data:
            ws.append(row)

        # Formatiere Prozente
        for row in ws.iter_rows(min_row=2, max_row=3, min_col=2, max_col=4):
            for cell in row:
                cell.number_format = '0.00%'

        # Füge Ergebnistabelle hinzu
        ws.append([])
        ws.append(['Test Results'])
        ws.append(['Relative Change', result['relative_change']])
        ws.append(['Minimum Effect Size', result['min_effect_size']])
        ws.append(['Confidence Level', result['confidence']])
        ws.append(['p-value', result.get('p_value', 'N/A')])
        ws.append(['Statistically Significant Improvement', 'Yes' if result['significant_improvement'] else 'No'])

        # Formatiere Prozente in der Ergebnistabelle
        ws.cell(row=7, column=2).number_format = '0.00%'
        ws.cell(row=8, column=2).number_format = '0.00%'
        ws.cell(row=9, column=2).number_format = '0.00%'

        # Erstelle eine zusätzliche Tabelle für Think-Cell
        ws.append([])
        ws.append(['Think-Cell Data'])

        # Spaltenformat, das für Think-Cell gut funktioniert
        tc_data = [
            ['Version', 'Rate', 'CI Min', 'CI Max'],
            [result['A_name'], result['A_rate'], result['A_rate'] - result['A_ci_lower'],
             result['A_ci_upper'] - result['A_rate']],
            [result['B_name'], result['B_rate'], result['B_rate'] - result['B_ci_lower'],
             result['B_ci_upper'] - result['B_rate']]
        ]

        # Füge Think-Cell Daten hinzu
        row_offset = 13
        for i, row in enumerate(tc_data):
            for j, value in enumerate(row):
                ws.cell(row=row_offset + i, column=j + 1, value=value)
                if i > 0 and j > 0:
                    ws.cell(row=row_offset + i, column=j + 1).number_format = '0.00%'

        # Erstelle ein separates Blatt für die Normalverteilungsdaten in hoher Auflösung
        norm_ws = wb.create_sheet(title=f"{metric}_Normal")

        # Berechne Standardfehler für jede Version
        # Für Proportionen: Standardfehler = sqrt(p * (1-p) / n)
        se_a = np.sqrt(result['A_rate'] * (1 - result['A_rate']) / result.get('A_sample_size', 10000))
        se_b = np.sqrt(result['B_rate'] * (1 - result['B_rate']) / result.get('B_sample_size', 10000))

        # Erstelle Normalverteilungsdaten für beide Versionen mit sehr hoher Auflösung
        # und Entfernung der extremen Randbereiche (nur 1% des Maximums oder höher)
        x_a, y_a = create_normal_distribution_data(result['A_rate'], se_a, num_points=500, cutoff_threshold=0.01)
        x_b, y_b = create_normal_distribution_data(result['B_rate'], se_b, num_points=500, cutoff_threshold=0.01)

        # Erstelle Überschriften für die Normalverteilungsdaten
        norm_ws.append(['X_A', 'Y_A', 'X_B', 'Y_B'])

        # Da die Kurven unterschiedliche Längen haben können, ermitteln wir das Maximum
        max_len = max(len(x_a), len(x_b))

        # Erstelle Arrays mit entsprechender Länge und fülle fehlende Werte mit None
        x_a_padded = list(x_a) + [None] * (max_len - len(x_a))
        y_a_padded = list(y_a) + [None] * (max_len - len(y_a))
        x_b_padded = list(x_b) + [None] * (max_len - len(x_b))
        y_b_padded = list(y_b) + [None] * (max_len - len(y_b))

        # Füge Normalverteilungsdaten hinzu
        for i in range(max_len):
            norm_ws.append([
                None if i >= len(x_a) else float(x_a[i]),
                None if i >= len(y_a) else float(y_a[i]),
                None if i >= len(x_b) else float(x_b[i]),
                None if i >= len(y_b) else float(y_b[i])
            ])

        # Formatiere Prozente für X-Werte
        for row in norm_ws.iter_rows(min_row=2, max_row=len(x_a) + 1, min_col=1, max_col=1):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = '0.00%'

        for row in norm_ws.iter_rows(min_row=2, max_row=len(x_b) + 1, min_col=3, max_col=3):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = '0.00%'

        # Füge zusätzliche Informationen für die Normalverteilungskurven hinzu
        norm_ws.append([])
        norm_ws.append(['Version', 'Mean', 'Standard Error', 'CI Lower', 'CI Upper'])
        norm_ws.append([result['A_name'], result['A_rate'], se_a, result['A_ci_lower'], result['A_ci_upper']])
        norm_ws.append([result['B_name'], result['B_rate'], se_b, result['B_ci_lower'], result['B_ci_upper']])

        # Formatiere Prozente
        row_offset = max_len + 3
        for row in norm_ws.iter_rows(min_row=row_offset + 1, max_row=row_offset + 2, min_col=2, max_col=5):
            for cell in row:
                cell.number_format = '0.00%'

    # Erstelle eine separate Datei speziell für Think-Cell mit besserer Formatierung
    curves_wb = Workbook()

    # Für jede Metrik
    for result in results:
        metric = result['metric']

        # Berechne Standardfehler für jede Version
        se_a = np.sqrt(result['A_rate'] * (1 - result['A_rate']) / result.get('A_sample_size', 10000))
        se_b = np.sqrt(result['B_rate'] * (1 - result['B_rate']) / result.get('B_sample_size', 10000))

        # Erstelle gefilterte Normalverteilungsdaten (ohne die extremen Ränder)
        x_a, y_a = create_normal_distribution_data(result['A_rate'], se_a, num_points=800, cutoff_threshold=0.005)
        x_b, y_b = create_normal_distribution_data(result['B_rate'], se_b, num_points=800, cutoff_threshold=0.005)

        # Erstelle ein Blatt für jede Metrik
        curves_ws = curves_wb.create_sheet(title=metric)

        # Spalte A und B für die erste Kurve
        curves_ws.append(['Kurve A - ' + result['A_name']])
        curves_ws.append(['X', 'Y'])

        # Füge Daten für Kurve A hinzu
        for i in range(len(x_a)):
            curves_ws.append([float(x_a[i]), float(y_a[i])])

        # Formatiere X-Werte als Prozente
        for row in curves_ws.iter_rows(min_row=3, max_row=len(x_a) + 2, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = '0.00%'

        # Spalte D und E für die zweite Kurve (mit Leerzeile dazwischen)
        col_offset = 3  # Spalte D
        curves_ws.cell(row=1, column=col_offset, value='Kurve B - ' + result['B_name'])
        curves_ws.cell(row=2, column=col_offset, value='X')
        curves_ws.cell(row=2, column=col_offset + 1, value='Y')

        # Füge Daten für Kurve B hinzu
        for i in range(len(x_b)):
            curves_ws.cell(row=i + 3, column=col_offset, value=float(x_b[i]))
            curves_ws.cell(row=i + 3, column=col_offset + 1, value=float(y_b[i]))

        # Formatiere X-Werte als Prozente
        for row in curves_ws.iter_rows(min_row=3, max_row=len(x_b) + 2, min_col=col_offset, max_col=col_offset):
            for cell in row:
                cell.number_format = '0.00%'

        # Füge statistische Informationen hinzu
        info_row = max(len(x_a), len(x_b)) + 5
        curves_ws.cell(row=info_row, column=1, value='Statistik')
        curves_ws.cell(row=info_row + 1, column=1, value='Metrik')
        curves_ws.cell(row=info_row + 1, column=2, value=metric)

        curves_ws.cell(row=info_row + 2, column=1, value='Version')
        curves_ws.cell(row=info_row + 2, column=2, value='Wert')
        curves_ws.cell(row=info_row + 2, column=3, value='CI Unten')
        curves_ws.cell(row=info_row + 2, column=4, value='CI Oben')

        curves_ws.cell(row=info_row + 3, column=1, value=result['A_name'])
        curves_ws.cell(row=info_row + 3, column=2, value=result['A_rate'])
        curves_ws.cell(row=info_row + 3, column=3, value=result['A_ci_lower'])
        curves_ws.cell(row=info_row + 3, column=4, value=result['A_ci_upper'])

        curves_ws.cell(row=info_row + 4, column=1, value=result['B_name'])
        curves_ws.cell(row=info_row + 4, column=2, value=result['B_rate'])
        curves_ws.cell(row=info_row + 4, column=3, value=result['B_ci_lower'])
        curves_ws.cell(row=info_row + 4, column=4, value=result['B_ci_upper'])

        # Formatiere Werte als Prozente
        for row in curves_ws.iter_rows(min_row=info_row + 3, max_row=info_row + 4, min_col=2, max_col=4):
            for cell in row:
                cell.number_format = '0.00%'

    # Entferne die standardmäßigen Arbeitsblätter
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    if 'Sheet' in curves_wb.sheetnames:
        del curves_wb['Sheet']

    # Speichere die Dateien
    wb.save(filename)
    curves_wb.save("kurven_" + filename)

    print(f"Ergebnisse wurden in {filename} gespeichert.")
    print(f"Optimierte Kurvendaten wurden in kurven_{filename} gespeichert.")


def calculate_distribution_overlap(mean1, std1, mean2, std2):
    """
    Berechnet den Überlappungsgrad zweier Normalverteilungen.
    Gibt einen Wert zwischen 0 (keine Überlappung) und 1 (vollständige Überlappung) zurück.
    """
    # Vereinfachte Überlappungsberechnung
    if mean1 > mean2:
        mean1, mean2 = mean2, mean1
        std1, std2 = std2, std1

    d = (mean2 - mean1) / np.sqrt((std1 ** 2 + std2 ** 2) / 2)  # Cohen's d
    overlap = 2 * stats.norm.cdf(-abs(d) / 2)

    return overlap


def interpret_overlap(overlap):
    """
    Interpretiert den Überlappungsgrad zwischen zwei Verteilungen.
    """
    if overlap < 0.2:
        return "Minimal Overlap - Strong Evidence for Difference"
    elif overlap < 0.4:
        return "Low Overlap - Good Evidence for Difference"
    elif overlap < 0.7:
        return "Moderate Overlap - Some Evidence for Difference"
    else:
        return "High Overlap - Weak Evidence for Difference"


def plot_results(results):
    """
    Erstellt Plots für die Ergebnisse (für Visualisierung im Python-Skript).
    """
    for result in results:
        metric = result['metric']

        # Erstelle Figur und Achsen für Balkendiagramm mit Konfidenzintervallen
        fig, ax = plt.subplots(figsize=(10, 6))

        # Datenpunkte
        versions = [result['A_name'], result['B_name']]
        rates = [result['A_rate'], result['B_rate']]
        lower_errors = [result['A_rate'] - result['A_ci_lower'], result['B_rate'] - result['B_ci_lower']]
        upper_errors = [result['A_ci_upper'] - result['A_rate'], result['B_ci_upper'] - result['B_rate']]

        # Erstelle asymmetrische Fehlerbalken
        ax.errorbar(versions, rates, yerr=[lower_errors, upper_errors], fmt='o', capsize=5,
                    ecolor='black', markerfacecolor='blue', markersize=8)

        # Beschriftungen und Titel
        ax.set_ylabel(f'{metric} Rate')
        ax.set_title(f'A/B Test Results: {metric}\n{result["confidence"] * 100}% Confidence Intervals')

        # Füge Informationen zur relativen Änderung hinzu
        change_text = f'Relative Change: {result["relative_change"] * 100:.2f}%\n'
        change_text += f'Minimum Effect Size: {result["min_effect_size"] * 100:.2f}%\n'
        change_text += f'Statistically Significant: {"Yes" if result["significant_improvement"] else "No"}'

        ax.text(0.02, 0.02, change_text, transform=ax.transAxes,
                bbox=dict(facecolor='white', alpha=0.8))

        # Anzeige des Plots
        plt.tight_layout()
        plt.savefig(f'{metric}_result.png')
        plt.close()

        # Erstelle Normalverteilungskurven
        fig, ax = plt.subplots(figsize=(12, 6))

        # Berechne Standardfehler für jede Version
        se_a = np.sqrt(result['A_rate'] * (1 - result['A_rate']) / result.get('A_sample_size', 10000))
        se_b = np.sqrt(result['B_rate'] * (1 - result['B_rate']) / result.get('B_sample_size', 10000))

        # Erzeuge x-Werte für die Kurven
        x_min = min(result['A_rate'], result['B_rate']) - 4 * max(se_a, se_b)
        x_max = max(result['A_rate'], result['B_rate']) + 4 * max(se_a, se_b)
        x = np.linspace(x_min, x_max, 1000)

        # Erzeuge Normalverteilungskurven
        pdf_a = stats.norm.pdf(x, result['A_rate'], se_a)
        pdf_b = stats.norm.pdf(x, result['B_rate'], se_b)

        # Plot Normalverteilungskurven
        ax.plot(x, pdf_a, 'b-', linewidth=2, label=f"{result['A_name']} ({result['A_rate'] * 100:.2f}%)")
        ax.plot(x, pdf_b, 'r-', linewidth=2, label=f"{result['B_name']} ({result['B_rate'] * 100:.2f}%)")

        # Fülle Konfidenzintervalle
        ax.fill_between(x, 0, pdf_a, where=(x >= result['A_ci_lower']) & (x <= result['A_ci_upper']),
                        color='blue', alpha=0.3, label=f"{result['confidence'] * 100:.0f}% CI (A)")
        ax.fill_between(x, 0, pdf_b, where=(x >= result['B_ci_lower']) & (x <= result['B_ci_upper']),
                        color='red', alpha=0.3, label=f"{result['confidence'] * 100:.0f}% CI (B)")

        # Beschriftungen und Titel
        ax.set_xlabel(f'{metric} Rate')
        ax.set_ylabel('Probability Density')
        ax.set_title(f'Normal Distribution of {metric} Rates\np-value: {result["p_value"]:.4f}')

        # Füge Legende hinzu
        ax.legend()

        # Anzeige des Plots
        plt.tight_layout()
        plt.savefig(f'{metric}_normal_dist.png')
        plt.close()


def create_continuous_curve_data(results, filename="continuous_curves.xlsx"):
    """
    Erstellt eine Excel-Datei mit kontinuierlichen Y-Werten für Normalverteilungskurven,
    wobei beide Kurven an den exakt gleichen X-Positionen abgetastet werden.

    Args:
        results: Liste der A/B-Test-Ergebnisse
        filename: Name der zu erstellenden Excel-Datei
    """
    wb = Workbook()

    # Für jede Metrik
    for result in results:
        metric = result['metric']

        # Berechne Standardfehler für jede Version
        se_a = np.sqrt(result['A_rate'] * (1 - result['A_rate']) / result.get('A_sample_size', 10000))
        se_b = np.sqrt(result['B_rate'] * (1 - result['B_rate']) / result.get('B_sample_size', 10000))

        # Für eine korrekte Visualisierung müssen wir ein gemeinsames X-Intervall mit gleichem Abstand verwenden
        # Wir müssen die X-Werte so wählen, dass beide Kurven vollständig abgedeckt sind

        # Finde ein X-Intervall, das beide Kurven komplett abdeckt
        joint_min = min(result['A_rate'] - 4 * se_a, result['B_rate'] - 4 * se_b)
        joint_max = max(result['A_rate'] + 4 * se_a, result['B_rate'] + 4 * se_b)

        # Erstelle gleichmäßig verteilte X-Werte im gemeinsamen Bereich
        num_points = 300  # Genügend Punkte für eine glatte Kurve
        x_values = np.linspace(joint_min, joint_max, num_points)

        # Berechne die Y-Werte für beide Kurven an den exakt gleichen X-Positionen
        y_a = stats.norm.pdf(x_values, result['A_rate'], se_a)
        y_b = stats.norm.pdf(x_values, result['B_rate'], se_b)

        # Normalisiere beide Kurven auf den gleichen Maximalwert (1.0)
        y_a_normalized = y_a / np.max(y_a)
        y_b_normalized = y_b / np.max(y_b)

        # Erstelle ein Blatt für die Vergleichskurven
        ws = wb.create_sheet(title=f"{metric}_Comparison")
        ws.append([f"{metric} - Normalisierte Kurven mit gemeinsamer X-Achse"])
        ws.append(["X-Wert", "A-Kurve", "B-Kurve"])

        # Füge alle Datenpunkte hinzu
        for i in range(len(x_values)):
            ws.append([float(x_values[i]), float(y_a_normalized[i]), float(y_b_normalized[i])])

        # Formatiere X-Werte als Prozente
        for row in ws.iter_rows(min_row=3, max_row=len(x_values) + 2, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = '0.00%'

        # Erstelle ein Blatt speziell für Think-Cell (nur Y-Werte, genau an den richtigen Positionen)
        tc_ws = wb.create_sheet(title=f"{metric}_ThinkCell")
        tc_ws.append([f"{metric} - Nur Y-Werte in relativen Positionen"])
        tc_ws.append(["Y-Werte A", "Y-Werte B"])

        # Entferne extrem niedrige Werte, die unter einem Schwellenwert liegen
        # Der Schwellenwert sollte für beide Kurven einheitlich sein
        threshold = 0.005  # 0.5% des Maximalwerts
        mask = (y_a_normalized > threshold) | (y_b_normalized > threshold)

        # Wenn die Maske zu viele Werte herausfiltert, verwende eine kleinere Schwelle
        if np.sum(mask) < 50:
            threshold = 0.001  # 0.1% des Maximalwerts
            mask = (y_a_normalized > threshold) | (y_b_normalized > threshold)

        filtered_x = x_values[mask]
        filtered_y_a = y_a_normalized[mask]
        filtered_y_b = y_b_normalized[mask]

        # Füge die gefilterten Y-Werte hinzu
        for i in range(len(filtered_y_a)):
            tc_ws.append([float(filtered_y_a[i]), float(filtered_y_b[i])])

        # Erstelle ein separates Blatt mit den Y-Werten in Prozent (für bessere Lesbarkeit)
        percent_ws = wb.create_sheet(title=f"{metric}_Percent")
        percent_ws.append([f"{metric} - Kurven in Prozent"])
        percent_ws.append(["X-Wert", "A (%)", "B (%)"])

        # Berechne die Prozentsätze: Die Kurve A sollte bei A_rate zentriert sein
        # und die Kurve B bei B_rate
        for i in range(len(filtered_x)):
            percent_ws.append([
                f"{filtered_x[i] * 100:.2f}%",
                f"{filtered_y_a[i] * 100:.2f}%",
                f"{filtered_y_b[i] * 100:.2f}%"
            ])

        # Erstelle ein Datenblatt mit gleichmäßig verteilten Y-Werten (200 Punkte)
        # Dieses ist ideal für Think-Cell, da es exakt gleichmäßige Abstände garantiert
        smooth_ws = wb.create_sheet(title=f"{metric}_200Points")
        smooth_ws.append([f"{metric} - 200 gleichmäßige Punkte"])
        smooth_ws.append(["Y-Werte A", "Y-Werte B"])

        # Berechne genau 200 gleichmäßig verteilte Punkte
        indices = np.round(np.linspace(0, len(filtered_y_a) - 1, 200)).astype(int)
        smooth_y_a = filtered_y_a[indices]
        smooth_y_b = filtered_y_b[indices]

        # Füge die 200 gleichmäßigen Punkte hinzu
        for i in range(len(smooth_y_a)):
            smooth_ws.append([float(smooth_y_a[i]), float(smooth_y_b[i])])

    # Entferne das standardmäßige Arbeitsblatt
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Speichere die Datei
    wb.save(filename)

    print(f"Optimierte Kurvendaten wurden in {filename} gespeichert.")
    import pandas as pd


import numpy as np
import scipy.stats as stats
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import os


def calculate_confidence_interval(successes, total, confidence=0.95):
    """
    Berechnet das Konfidenzintervall für eine Proportion mit der Wilson-Score-Methode.
    """
    if total == 0:
        return 0, 0, 0

    proportion = successes / total
    z = stats.norm.ppf((1 + confidence) / 2)

    # Wilson Score Interval
    denominator = 1 + z ** 2 / total
    center = (proportion + z ** 2 / (2 * total)) / denominator
    margin = z * np.sqrt(proportion * (1 - proportion) / total + z ** 2 / (4 * total ** 2)) / denominator

    lower_bound = max(0, center - margin)
    upper_bound = min(1, center + margin)

    return proportion, lower_bound, upper_bound


def run_ab_test(data_a, data_b, metric_name, min_effect_size=0.05, confidence=0.95):
    """
    Führt einen A/B-Test durch und gibt die Ergebnisse zurück.
    data_a: Dictionary mit 'success' und 'total' für Variante A
    data_b: Dictionary mit 'success' und 'total' für Variante B
    metric_name: Name der Metrik (z.B. 'CTR' oder 'Bounce Rate')
    min_effect_size: Minimale Effektgröße, die als signifikant betrachtet wird
    confidence: Konfidenzniveau (Standard: 0.95)
    """
    # Berechne Konversionsraten und Konfidenzintervalle
    rate_a, ci_lower_a, ci_upper_a = calculate_confidence_interval(data_a['success'], data_a['total'], confidence)
    rate_b, ci_lower_b, ci_upper_b = calculate_confidence_interval(data_b['success'], data_b['total'], confidence)

    # Berechne relative Änderung (für CTR positiv, für Bounce Rate negativ)
    if metric_name == "CTR":
        relative_change = (rate_b - rate_a) / rate_a
        significant_improvement = rate_b > rate_a * (1 + min_effect_size) and ci_lower_b > ci_upper_a
    elif metric_name == "Bounce Rate":
        relative_change = (rate_a - rate_b) / rate_a  # Für Bounce Rate ist eine Reduktion besser
        significant_improvement = rate_b < rate_a * (1 - min_effect_size) and ci_upper_b < ci_lower_a

    # Standardabweichung für beide Varianten berechnen
    # Für Proportionen: Standardabweichung = sqrt(p * (1-p))
    std_dev_a = np.sqrt(rate_a * (1 - rate_a))
    std_dev_b = np.sqrt(rate_b * (1 - rate_b))

    # Z-Score für den Unterschied berechnen
    pooled_std = np.sqrt(std_dev_a ** 2 / data_a['total'] + std_dev_b ** 2 / data_b['total'])
    z_score = (rate_b - rate_a) / pooled_std if pooled_std > 0 else 0

    # p-Wert berechnen
    p_value = 2 * (1 - stats.norm.cdf(abs(z_score)))

    # Erstelle Ergebnisdictionary
    result = {
        'metric': metric_name,
        'A_name': "Applications Open Now",
        'B_name': "Applications Closing Soon",
        'A_rate': rate_a,
        'A_ci_lower': ci_lower_a,
        'A_ci_upper': ci_upper_a,
        'A_sample_size': data_a['total'],
        'B_rate': rate_b,
        'B_ci_lower': ci_lower_b,
        'B_ci_upper': ci_upper_b,
        'B_sample_size': data_b['total'],
        'relative_change': relative_change,
        'min_effect_size': min_effect_size,
        'confidence': confidence,
        'z_score': z_score,
        'p_value': p_value,
        'significant_improvement': significant_improvement
    }

    return result


def create_normal_distribution_data(mean, std_dev, num_points=500, cutoff_threshold=0.01):
    """
    Erstellt Datenpunkte für eine Normalverteilung mit sehr feinen Abstufungen.
    Schneidet die Verteilung bei einem bestimmten Schwellenwert ab, um nur relevante Bereiche einzuschließen.

    Args:
        mean: Mittelwert der Normalverteilung
        std_dev: Standardabweichung der Normalverteilung
        num_points: Anzahl der zu erzeugenden Datenpunkte (höher = feinere Kurve)
        cutoff_threshold: Schwellenwert für y-Werte (alle Werte unter diesem Schwellenwert werden entfernt)

    Returns:
        Tuple aus zwei Listen (x-Werte, y-Werte) für die relevanten Bereiche der Kurve
    """
    # Erstelle zunächst einen breiten Bereich
    range_factor = 4
    x_min = mean - range_factor * std_dev
    x_max = mean + range_factor * std_dev

    # Erzeuge feine Datenpunkte
    x_values = np.linspace(x_min, x_max, num_points)
    y_values = stats.norm.pdf(x_values, mean, std_dev)

    # Finde den maximalen y-Wert
    y_max = np.max(y_values)

    # Berechne den absoluten Schwellenwert basierend auf dem relativen Schwellenwert
    abs_threshold = y_max * cutoff_threshold

    # Filtere x- und y-Werte, wo y über dem Schwellenwert liegt
    mask = y_values > abs_threshold
    filtered_x = x_values[mask]
    filtered_y = y_values[mask]

    return filtered_x, filtered_y


def create_excel_for_think_cell(results, filename="ab_test_results.xlsx"):
    """
    Erstellt eine Excel-Datei mit den Ergebnissen, die für Think-Cell optimiert ist.
    Enthält zusätzlich Daten für Normalverteilungskurven mit sehr feinen Abstufungen.
    """
    wb = Workbook()

    # Erstelle ein Blatt für jede Metrik
    for result in results:
        metric = result['metric']
        ws = wb.create_sheet(title=metric)

        # Erstelle Daten für die Kurven
        data = [
            ['Version', 'Rate', 'Lower CI', 'Upper CI'],
            [result['A_name'], result['A_rate'], result['A_ci_lower'], result['A_ci_upper']],
            [result['B_name'], result['B_rate'], result['B_ci_lower'], result['B_ci_upper']]
        ]

        # Füge Daten zum Arbeitsblatt hinzu
        for row in data:
            ws.append(row)

        # Formatiere Prozente
        for row in ws.iter_rows(min_row=2, max_row=3, min_col=2, max_col=4):
            for cell in row:
                cell.number_format = '0.00%'

        # Füge Ergebnistabelle hinzu
        ws.append([])
        ws.append(['Test Results'])
        ws.append(['Relative Change', result['relative_change']])
        ws.append(['Minimum Effect Size', result['min_effect_size']])
        ws.append(['Confidence Level', result['confidence']])
        ws.append(['p-value', result.get('p_value', 'N/A')])
        ws.append(['Statistically Significant Improvement', 'Yes' if result['significant_improvement'] else 'No'])

        # Formatiere Prozente in der Ergebnistabelle
        ws.cell(row=7, column=2).number_format = '0.00%'
        ws.cell(row=8, column=2).number_format = '0.00%'
        ws.cell(row=9, column=2).number_format = '0.00%'

        # Erstelle eine zusätzliche Tabelle für Think-Cell
        ws.append([])
        ws.append(['Think-Cell Data'])

        # Spaltenformat, das für Think-Cell gut funktioniert
        tc_data = [
            ['Version', 'Rate', 'CI Min', 'CI Max'],
            [result['A_name'], result['A_rate'], result['A_rate'] - result['A_ci_lower'],
             result['A_ci_upper'] - result['A_rate']],
            [result['B_name'], result['B_rate'], result['B_rate'] - result['B_ci_lower'],
             result['B_ci_upper'] - result['B_rate']]
        ]

        # Füge Think-Cell Daten hinzu
        row_offset = 13
        for i, row in enumerate(tc_data):
            for j, value in enumerate(row):
                ws.cell(row=row_offset + i, column=j + 1, value=value)
                if i > 0 and j > 0:
                    ws.cell(row=row_offset + i, column=j + 1).number_format = '0.00%'

        # Erstelle ein separates Blatt für die Normalverteilungsdaten in hoher Auflösung
        norm_ws = wb.create_sheet(title=f"{metric}_Normal")

        # Berechne Standardfehler für jede Version
        # Für Proportionen: Standardfehler = sqrt(p * (1-p) / n)
        se_a = np.sqrt(result['A_rate'] * (1 - result['A_rate']) / result.get('A_sample_size', 10000))
        se_b = np.sqrt(result['B_rate'] * (1 - result['B_rate']) / result.get('B_sample_size', 10000))

        # Erstelle Normalverteilungsdaten für beide Versionen mit sehr hoher Auflösung
        # und Entfernung der extremen Randbereiche (nur 1% des Maximums oder höher)
        x_a, y_a = create_normal_distribution_data(result['A_rate'], se_a, num_points=500, cutoff_threshold=0.01)
        x_b, y_b = create_normal_distribution_data(result['B_rate'], se_b, num_points=500, cutoff_threshold=0.01)

        # Erstelle Überschriften für die Normalverteilungsdaten
        norm_ws.append(['X_A', 'Y_A', 'X_B', 'Y_B'])

        # Da die Kurven unterschiedliche Längen haben können, ermitteln wir das Maximum
        max_len = max(len(x_a), len(x_b))

        # Erstelle Arrays mit entsprechender Länge und fülle fehlende Werte mit None
        x_a_padded = list(x_a) + [None] * (max_len - len(x_a))
        y_a_padded = list(y_a) + [None] * (max_len - len(y_a))
        x_b_padded = list(x_b) + [None] * (max_len - len(x_b))
        y_b_padded = list(y_b) + [None] * (max_len - len(y_b))

        # Füge Normalverteilungsdaten hinzu
        for i in range(max_len):
            norm_ws.append([
                None if i >= len(x_a) else float(x_a[i]),
                None if i >= len(y_a) else float(y_a[i]),
                None if i >= len(x_b) else float(x_b[i]),
                None if i >= len(y_b) else float(y_b[i])
            ])

        # Formatiere Prozente für X-Werte
        for row in norm_ws.iter_rows(min_row=2, max_row=len(x_a) + 1, min_col=1, max_col=1):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = '0.00%'

        for row in norm_ws.iter_rows(min_row=2, max_row=len(x_b) + 1, min_col=3, max_col=3):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = '0.00%'

        # Füge zusätzliche Informationen für die Normalverteilungskurven hinzu
        norm_ws.append([])
        norm_ws.append(['Version', 'Mean', 'Standard Error', 'CI Lower', 'CI Upper'])
        norm_ws.append([result['A_name'], result['A_rate'], se_a, result['A_ci_lower'], result['A_ci_upper']])
        norm_ws.append([result['B_name'], result['B_rate'], se_b, result['B_ci_lower'], result['B_ci_upper']])

        # Formatiere Prozente
        row_offset = max_len + 3
        for row in norm_ws.iter_rows(min_row=row_offset + 1, max_row=row_offset + 2, min_col=2, max_col=5):
            for cell in row:
                cell.number_format = '0.00%'

    # Erstelle eine separate Datei speziell für Think-Cell mit besserer Formatierung
    curves_wb = Workbook()

    # Für jede Metrik
    for result in results:
        metric = result['metric']

        # Berechne Standardfehler für jede Version
        se_a = np.sqrt(result['A_rate'] * (1 - result['A_rate']) / result.get('A_sample_size', 10000))
        se_b = np.sqrt(result['B_rate'] * (1 - result['B_rate']) / result.get('B_sample_size', 10000))

        # Erstelle gefilterte Normalverteilungsdaten (ohne die extremen Ränder)
        x_a, y_a = create_normal_distribution_data(result['A_rate'], se_a, num_points=800, cutoff_threshold=0.005)
        x_b, y_b = create_normal_distribution_data(result['B_rate'], se_b, num_points=800, cutoff_threshold=0.005)

        # Erstelle ein Blatt für jede Metrik
        curves_ws = curves_wb.create_sheet(title=metric)

        # Spalte A und B für die erste Kurve
        curves_ws.append(['Kurve A - ' + result['A_name']])
        curves_ws.append(['X', 'Y'])

        # Füge Daten für Kurve A hinzu
        for i in range(len(x_a)):
            curves_ws.append([float(x_a[i]), float(y_a[i])])

        # Formatiere X-Werte als Prozente
        for row in curves_ws.iter_rows(min_row=3, max_row=len(x_a) + 2, min_col=1, max_col=1):
            for cell in row:
                cell.number_format = '0.00%'

        # Spalte D und E für die zweite Kurve (mit Leerzeile dazwischen)
        col_offset = 3  # Spalte D
        curves_ws.cell(row=1, column=col_offset, value='Kurve B - ' + result['B_name'])
        curves_ws.cell(row=2, column=col_offset, value='X')
        curves_ws.cell(row=2, column=col_offset + 1, value='Y')

        # Füge Daten für Kurve B hinzu
        for i in range(len(x_b)):
            curves_ws.cell(row=i + 3, column=col_offset, value=float(x_b[i]))
            curves_ws.cell(row=i + 3, column=col_offset + 1, value=float(y_b[i]))

        # Formatiere X-Werte als Prozente
        for row in curves_ws.iter_rows(min_row=3, max_row=len(x_b) + 2, min_col=col_offset, max_col=col_offset):
            for cell in row:
                cell.number_format = '0.00%'

        # Füge statistische Informationen hinzu
        info_row = max(len(x_a), len(x_b)) + 5
        curves_ws.cell(row=info_row, column=1, value='Statistik')
        curves_ws.cell(row=info_row + 1, column=1, value='Metrik')
        curves_ws.cell(row=info_row + 1, column=2, value=metric)

        curves_ws.cell(row=info_row + 2, column=1, value='Version')
        curves_ws.cell(row=info_row + 2, column=2, value='Wert')
        curves_ws.cell(row=info_row + 2, column=3, value='CI Unten')
        curves_ws.cell(row=info_row + 2, column=4, value='CI Oben')

        curves_ws.cell(row=info_row + 3, column=1, value=result['A_name'])
        curves_ws.cell(row=info_row + 3, column=2, value=result['A_rate'])
        curves_ws.cell(row=info_row + 3, column=3, value=result['A_ci_lower'])
        curves_ws.cell(row=info_row + 3, column=4, value=result['A_ci_upper'])

        curves_ws.cell(row=info_row + 4, column=1, value=result['B_name'])
        curves_ws.cell(row=info_row + 4, column=2, value=result['B_rate'])
        curves_ws.cell(row=info_row + 4, column=3, value=result['B_ci_lower'])
        curves_ws.cell(row=info_row + 4, column=4, value=result['B_ci_upper'])

        # Formatiere Werte als Prozente
        for row in curves_ws.iter_rows(min_row=info_row + 3, max_row=info_row + 4, min_col=2, max_col=4):
            for cell in row:
                cell.number_format = '0.00%'

    # Entferne die standardmäßigen Arbeitsblätter
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    if 'Sheet' in curves_wb.sheetnames:
        del curves_wb['Sheet']

    # Speichere die Dateien
    wb.save(filename)
    curves_wb.save("kurven_" + filename)

    print(f"Ergebnisse wurden in {filename} gespeichert.")
    print(f"Optimierte Kurvendaten wurden in kurven_{filename} gespeichert.")


def calculate_distribution_overlap(mean1, std1, mean2, std2):
    """
    Berechnet den Überlappungsgrad zweier Normalverteilungen.
    Gibt einen Wert zwischen 0 (keine Überlappung) und 1 (vollständige Überlappung) zurück.
    """
    # Vereinfachte Überlappungsberechnung
    if mean1 > mean2:
        mean1, mean2 = mean2, mean1
        std1, std2 = std2, std1

    d = (mean2 - mean1) / np.sqrt((std1 ** 2 + std2 ** 2) / 2)  # Cohen's d
    overlap = 2 * stats.norm.cdf(-abs(d) / 2)

    return overlap


def interpret_overlap(overlap):
    """
    Interpretiert den Überlappungsgrad zwischen zwei Verteilungen.
    """
    if overlap < 0.2:
        return "Minimal Overlap - Strong Evidence for Difference"
    elif overlap < 0.4:
        return "Low Overlap - Good Evidence for Difference"
    elif overlap < 0.7:
        return "Moderate Overlap - Some Evidence for Difference"
    else:
        return "High Overlap - Weak Evidence for Difference"


def plot_results(results):
    """
    Erstellt Plots für die Ergebnisse (für Visualisierung im Python-Skript).
    """
    for result in results:
        metric = result['metric']

        # Erstelle Figur und Achsen für Balkendiagramm mit Konfidenzintervallen
        fig, ax = plt.subplots(figsize=(10, 6))

        # Datenpunkte
        versions = [result['A_name'], result['B_name']]
        rates = [result['A_rate'], result['B_rate']]
        lower_errors = [result['A_rate'] - result['A_ci_lower'], result['B_rate'] - result['B_ci_lower']]
        upper_errors = [result['A_ci_upper'] - result['A_rate'], result['B_ci_upper'] - result['B_rate']]

        # Erstelle asymmetrische Fehlerbalken
        ax.errorbar(versions, rates, yerr=[lower_errors, upper_errors], fmt='o', capsize=5,
                    ecolor='black', markerfacecolor='blue', markersize=8)

        # Beschriftungen und Titel
        ax.set_ylabel(f'{metric} Rate')
        ax.set_title(f'A/B Test Results: {metric}\n{result["confidence"] * 100}% Confidence Intervals')

        # Füge Informationen zur relativen Änderung hinzu
        change_text = f'Relative Change: {result["relative_change"] * 100:.2f}%\n'
        change_text += f'Minimum Effect Size: {result["min_effect_size"] * 100:.2f}%\n'
        change_text += f'Statistically Significant: {"Yes" if result["significant_improvement"] else "No"}'

        ax.text(0.02, 0.02, change_text, transform=ax.transAxes,
                bbox=dict(facecolor='white', alpha=0.8))

        # Anzeige des Plots
        plt.tight_layout()
        plt.savefig(f'{metric}_result.png')
        plt.close()

        # Erstelle Normalverteilungskurven
        fig, ax = plt.subplots(figsize=(12, 6))

        # Berechne Standardfehler für jede Version
        se_a = np.sqrt(result['A_rate'] * (1 - result['A_rate']) / result.get('A_sample_size', 10000))
        se_b = np.sqrt(result['B_rate'] * (1 - result['B_rate']) / result.get('B_sample_size', 10000))

        # Erzeuge x-Werte für die Kurven
        x_min = min(result['A_rate'], result['B_rate']) - 4 * max(se_a, se_b)
        x_max = max(result['A_rate'], result['B_rate']) + 4 * max(se_a, se_b)
        x = np.linspace(x_min, x_max, 1000)

        # Erzeuge Normalverteilungskurven
        pdf_a = stats.norm.pdf(x, result['A_rate'], se_a)
        pdf_b = stats.norm.pdf(x, result['B_rate'], se_b)

        # Plot Normalverteilungskurven
        ax.plot(x, pdf_a, 'b-', linewidth=2, label=f"{result['A_name']} ({result['A_rate'] * 100:.2f}%)")
        ax.plot(x, pdf_b, 'r-', linewidth=2, label=f"{result['B_name']} ({result['B_rate'] * 100:.2f}%)")

        # Fülle Konfidenzintervalle
        ax.fill_between(x, 0, pdf_a, where=(x >= result['A_ci_lower']) & (x <= result['A_ci_upper']),
                        color='blue', alpha=0.3, label=f"{result['confidence'] * 100:.0f}% CI (A)")
        ax.fill_between(x, 0, pdf_b, where=(x >= result['B_ci_lower']) & (x <= result['B_ci_upper']),
                        color='red', alpha=0.3, label=f"{result['confidence'] * 100:.0f}% CI (B)")

        # Beschriftungen und Titel
        ax.set_xlabel(f'{metric} Rate')
        ax.set_ylabel('Probability Density')
        ax.set_title(f'Normal Distribution of {metric} Rates\np-value: {result["p_value"]:.4f}')

        # Füge Legende hinzu
        ax.legend()

        # Anzeige des Plots
        plt.tight_layout()
        plt.savefig(f'{metric}_normal_dist.png')
        plt.close()


def create_thinkcell_table(results, filename="thinkcell_table.xlsx"):
    """
    Creates an Excel file with A/B test results in a format optimized for ThinkCell.

    Args:
        results: List of A/B test result dictionaries
        filename: Name of the Excel file to be created
    """
    from openpyxl import Workbook
    import pandas as pd

    # Create a new workbook
    wb = Workbook()

    # Delete the default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Create summary table for all metrics in one view
    summary_ws = wb.create_sheet(title="Summary")

    # Header row
    summary_header = ["Metric", "Variant", "Rate", "CI Lower", "CI Upper", "Sample Size",
                      "Relative Change", "p-value", "Significant"]
    summary_ws.append(summary_header)

    # Data rows - create a flattened format for easy comparison
    row_index = 2  # Start after header

    for result in results:
        metric = result['metric']

        # Add A variant data
        summary_ws.append([
            metric,
            result['A_name'],
            result['A_rate'],
            result['A_ci_lower'],
            result['A_ci_upper'],
            result['A_sample_size'],
            "",  # No relative change for A (baseline)
            "",  # No p-value for A (baseline)
            ""  # No significance for A (baseline)
        ])

        # Add B variant data with relative change and significance
        summary_ws.append([
            metric,
            result['B_name'],
            result['B_rate'],
            result['B_ci_lower'],
            result['B_ci_upper'],
            result['B_sample_size'],
            result['relative_change'],
            result['p_value'],
            "Yes" if result['significant_improvement'] else "No"
        ])

        # Format cells as percentages
        for row in range(row_index, row_index + 2):
            for col in range(3, 6):  # Rate, CI Lower, CI Upper
                summary_ws.cell(row=row, column=col).number_format = '0.00%'

            # Format relative change as percentage (only for B variant)
            if row == row_index + 1:
                summary_ws.cell(row=row, column=7).number_format = '0.00%'

        row_index += 2

    # Create a ThinkCell-optimized table
    tc_ws = wb.create_sheet(title="ThinkCell_Ready")

    # Create the table structure for bar charts with error bars
    tc_ws.append(["A/B Test Results - ThinkCell Format"])
    tc_ws.append([])

    tc_ws.append(["Bar Chart with Error Bars"])
    tc_ws.append(["Metric", "Variant", "Value", "CI Lower", "CI Upper", "Better?"])

    for result in results:
        tc_ws.append([
            result['metric'],
            result['A_name'],
            result['A_rate'],
            result['A_ci_lower'],
            result['A_ci_upper'],
            "No"
        ])

        # Check if B is better than A based on the metric type
        is_better = False
        if result['metric'] == "CTR":
            is_better = result['B_rate'] > result['A_rate'] and result['significant_improvement']
        elif result['metric'] == "Bounce Rate":
            is_better = result['B_rate'] < result['A_rate'] and result['significant_improvement']

        tc_ws.append([
            result['metric'],
            result['B_name'],
            result['B_rate'],
            result['B_ci_lower'],
            result['B_ci_upper'],
            "Yes" if is_better else "No"
        ])

    # Format percentage cells
    for row in tc_ws.iter_rows(min_row=5, max_row=5 + len(results) * 2 - 1, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = '0.00%'

    # Add waterfall chart data
    tc_ws.append([])
    tc_ws.append(["Waterfall Chart for Relative Changes"])
    tc_ws.append(["Metric", "Relative Change", "Significant"])

    for result in results:
        tc_ws.append([
            result['metric'],
            result['relative_change'],
            "Yes" if result['significant_improvement'] else "No"
        ])

        # Format relative change as percentage
        tc_ws.cell(row=tc_ws.max_row, column=2).number_format = '0.00%'

    # Add comparative normal distribution data
    tc_ws.append([])
    tc_ws.append(["Distribution Chart Comparison (Using normal distribution curves)"])
    tc_ws.append(["Metric", "Mean A", "Mean B", "SE A", "SE B", "p-value", "Significant"])

    for result in results:
        # Calculate standard errors
        se_a = (result['A_ci_upper'] - result['A_ci_lower']) / (2 * 1.96)  # Approximation using 95% CI
        se_b = (result['B_ci_upper'] - result['B_ci_lower']) / (2 * 1.96)

        tc_ws.append([
            result['metric'],
            result['A_rate'],
            result['B_rate'],
            se_a,
            se_b,
            result['p_value'],
            "Yes" if result['significant_improvement'] else "No"
        ])

        # Format percentage cells
        for col in range(2, 6):
            tc_ws.cell(row=tc_ws.max_row, column=col).number_format = '0.00%'

    # Create dedicated worksheet for each visualization type

    # 1. Bar chart data
    bar_ws = wb.create_sheet(title="Bar_Chart")
    bar_ws.append(["Variant", "CTR", "Bounce Rate"])
    bar_ws.append([results[0]['A_name'], results[0]['A_rate'], results[1]['A_rate']])
    bar_ws.append([results[0]['B_name'], results[0]['B_rate'], results[1]['B_rate']])

    # Format as percentages
    for row in bar_ws.iter_rows(min_row=2, max_row=3, min_col=2, max_col=3):
        for cell in row:
            cell.number_format = '0.00%'

    # 2. Error bar data (for ThinkCell's error bars)
    error_ws = wb.create_sheet(title="Error_Bars")
    error_ws.append(["Variant", "Metric", "Value", "Lower Error", "Upper Error"])

    for result in results:
        error_ws.append([
            result['A_name'],
            result['metric'],
            result['A_rate'],
            result['A_rate'] - result['A_ci_lower'],
            result['A_ci_upper'] - result['A_rate']
        ])
        error_ws.append([
            result['B_name'],
            result['metric'],
            result['B_rate'],
            result['B_rate'] - result['B_ci_lower'],
            result['B_ci_upper'] - result['B_rate']
        ])

    # Format as percentages
    for row in error_ws.iter_rows(min_row=2, max_row=1 + len(results) * 2, min_col=3, max_col=5):
        for cell in row:
            cell.number_format = '0.00%'

    # Save the workbook
    wb.save(filename)
    print(f"ThinkCell-ready table saved to {filename}")

    return filename


def create_thinkcell_boxplot_format(results, filename="thinkcell_boxplot_format.xlsx"):
    """
    Creates an Excel file with A/B test results in a boxplot-style format for ThinkCell.
    This format represents confidence intervals as if they were quartiles in a box plot.

    Args:
        results: List of A/B test result dictionaries
        filename: Name of the Excel file to be created
    """
    from openpyxl import Workbook

    # Create a new workbook
    wb = Workbook()

    # Delete the default sheet
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    # Create sheet for CTR boxplot-style data
    ctr_ws = wb.create_sheet(title="CTR_BoxPlot")

    # Create sheet for Bounce Rate boxplot-style data
    bounce_ws = wb.create_sheet(title="BounceRate_BoxPlot")

    # Get the CTR and Bounce Rate results
    ctr_result = next((r for r in results if r['metric'] == 'CTR'), None)
    bounce_result = next((r for r in results if r['metric'] == 'Bounce Rate'), None)

    # Process results sheets if they exist
    sheets_to_process = []
    if ctr_result:
        sheets_to_process.append((ctr_ws, ctr_result))
    if bounce_result:
        sheets_to_process.append((bounce_ws, bounce_result))

    # Header row for each sheet with the boxplot format
    for ws, result in sheets_to_process:
        # Create header row
        ws.append(['', 'Kategorie', result['A_name'], result['B_name']])

        # For each variant, calculate:
        # 1. Maximum (upper end of CI)
        # 2. Upper quartile (not used with CIs, but set to halfway between median and max for visualization)
        # 3. Median (point estimate)
        # 4. Lower quartile (not used with CIs, but set to halfway between min and median for visualization)
        # 5. Minimum (lower end of CI)

        # Row 1: Maximum (upper CI)
        ws.append(['Maximum (Whisker)', 'Serie', result['A_ci_upper'], result['B_ci_upper']])

        # Row 2: Upper "quartile" (set to midpoint between median and max)
        a_upper_mid = (result['A_rate'] + result['A_ci_upper']) / 2
        b_upper_mid = (result['B_rate'] + result['B_ci_upper']) / 2
        ws.append(['Upper quartile (Box)', '', a_upper_mid, b_upper_mid])

        # Row 3: Median (point estimate)
        ws.append(['Median', '', result['A_rate'], result['B_rate']])

        # Row 4: Lower "quartile" (set to midpoint between min and median)
        a_lower_mid = (result['A_rate'] + result['A_ci_lower']) / 2
        b_lower_mid = (result['B_rate'] + result['B_ci_lower']) / 2
        ws.append(['Lower quartile (Box)', '', a_lower_mid, b_lower_mid])

        # Row 5: Minimum (lower CI)
        ws.append(['Minimum (Whisker)', '', result['A_ci_lower'], result['B_ci_lower']])

        # Format all data cells as percentages
        for row in ws.iter_rows(min_row=2, max_row=6, min_col=3, max_col=4):
            for cell in row:
                cell.number_format = '0.00%'

    # Create a combined sheet with both metrics
    combined_ws = wb.create_sheet(title="Combined_BoxPlot")
    combined_ws.append(['', 'Kategorie', 'CTR - ' + ctr_result['A_name'], 'CTR - ' + ctr_result['B_name'],
                        'Bounce Rate - ' + bounce_result['A_name'], 'Bounce Rate - ' + bounce_result['B_name']])

    # Add the same structure but with all metrics in one table
    # Row 1: Maximum (upper CI)
    combined_ws.append(['Maximum (Whisker)', 'Serie',
                        ctr_result['A_ci_upper'], ctr_result['B_ci_upper'],
                        bounce_result['A_ci_upper'], bounce_result['B_ci_upper']])

    # Row 2: Upper "quartile"
    combined_ws.append(['Upper quartile (Box)', '',
                        (ctr_result['A_rate'] + ctr_result['A_ci_upper']) / 2,
                        (ctr_result['B_rate'] + ctr_result['B_ci_upper']) / 2,
                        (bounce_result['A_rate'] + bounce_result['A_ci_upper']) / 2,
                        (bounce_result['B_rate'] + bounce_result['B_ci_upper']) / 2])

    # Row 3: Median (point estimate)
    combined_ws.append(['Median', '',
                        ctr_result['A_rate'], ctr_result['B_rate'],
                        bounce_result['A_rate'], bounce_result['B_rate']])

    # Row 4: Lower "quartile"
    combined_ws.append(['Lower quartile (Box)', '',
                        (ctr_result['A_rate'] + ctr_result['A_ci_lower']) / 2,
                        (ctr_result['B_rate'] + ctr_result['B_ci_lower']) / 2,
                        (bounce_result['A_rate'] + bounce_result['A_ci_lower']) / 2,
                        (bounce_result['B_rate'] + bounce_result['B_ci_lower']) / 2])

    # Row 5: Minimum (lower CI)
    combined_ws.append(['Minimum (Whisker)', '',
                        ctr_result['A_ci_lower'], ctr_result['B_ci_lower'],
                        bounce_result['A_ci_lower'], bounce_result['B_ci_lower']])

    # Format all data cells as percentages
    for row in combined_ws.iter_rows(min_row=2, max_row=6, min_col=3, max_col=6):
        for cell in row:
            cell.number_format = '0.00%'

    # Save the workbook
    wb.save(filename)
    print(f"ThinkCell boxplot-format table saved to {filename}")

    return filename

# Hauptfunktion für die A/B-Test-Analyse
# Hier ist die aktualisierte Hauptfunktion mit Ausgabe der p-Werte und Alpha-Werte

# Hier ist die aktualisierte Hauptfunktion mit Ausgabe der p-Werte und Alpha-Werte

def main():
    # Reale Daten aus dem A/B-Test

    # Echte Daten für Click-Through Rate (CTR)
    # Da wir nur die Prozentwerte haben, nehmen wir eine fiktive Stichprobengröße an
    # Die Stichprobengröße beeinflusst die Breite der Konfidenzintervalle
    sample_size = 10000  # Angenommene Stichprobengröße (kann angepasst werden)

    ctr_a_percent = 24.88  # 24,88% CTR für "Applications Open Now"
    ctr_b_percent = 31.24  # 31,24% CTR für "Applications Closing Soon"

    # Umrechnung in absolute Zahlen
    ctr_data_a = {'success': int(sample_size * ctr_a_percent / 100), 'total': sample_size}
    ctr_data_b = {'success': int(sample_size * ctr_b_percent / 100), 'total': sample_size}

    # Echte Daten für Bounce Rate (niedriger ist besser)
    bounce_a_percent = 62.81  # 62,81% Bounce Rate für "Applications Open Now"
    bounce_b_percent = 65.00  # 65,00% Bounce Rate für "Applications Closing Soon"

    # Umrechnung in absolute Zahlen
    bounce_data_a = {'success': int(sample_size * bounce_a_percent / 100), 'total': sample_size}
    bounce_data_b = {'success': int(sample_size * bounce_b_percent / 100), 'total': sample_size}

    # Standardwert für Alpha (Signifikanzniveau) ist 0.05 (1 - Konfidenzniveau)
    alpha = 0.05

    # Führe Tests durch
    ctr_result = run_ab_test(ctr_data_a, ctr_data_b, "CTR", min_effect_size=0.05, confidence=1 - alpha)
    bounce_result = run_ab_test(bounce_data_a, bounce_data_b, "Bounce Rate", min_effect_size=0.05, confidence=1 - alpha)

    # Sammle alle Ergebnisse
    results = [ctr_result, bounce_result]

    # Erstelle Excel-Datei für Think-Cell
    create_excel_for_think_cell(results)

    # Erstelle optimierte ThinkCell-Tabellen mit der neuen Funktion
    thinkcell_file = create_thinkcell_table(results)

    # Erstelle Box Plot Format für ThinkCell
    boxplot_file = create_thinkcell_boxplot_format(results)

    # Erstelle Visualisierungen für die Überprüfung
    plot_results(results)

    # Output results to console in English
    for result in results:
        print(f"\nResults for {result['metric']}:")
        conf_level = result['confidence'] * 100
        print(f"Confidence level: {conf_level:.1f}%")
        print(f"A ({result['A_name']}): {result['A_rate']:.4f}")
        print(f"  {conf_level:.1f}% Confidence Interval: [{result['A_ci_lower']:.4f}, {result['A_ci_upper']:.4f}]")
        print(f"B ({result['B_name']}): {result['B_rate']:.4f}")
        print(f"  {conf_level:.1f}% Confidence Interval: [{result['B_ci_lower']:.4f}, {result['B_ci_upper']:.4f}]")
        print(f"Relative change: {result['relative_change'] * 100:.2f}%")
        print(f"p-value: {result['p_value']:.6f}")
        print(f"Alpha (significance level): {alpha:.6f}")
        print(f"Statistically significant: {'Yes' if result['p_value'] < alpha else 'No'}")
        print(
            f"Significant improvement (incl. minimum effect size): {'Yes' if result['significant_improvement'] else 'No'}")

    print(f"\nThinkCell visualization data has been saved to:")
    print(f"1. {thinkcell_file} - Optimized data tables for ThinkCell")
    print(f"2. {boxplot_file} - Box plot format for ThinkCell")
    print(f"3. ab_test_results.xlsx - Standard results data")
    print(f"4. kurven_ab_test_results.xlsx - Distribution curve data")


if __name__ == "__main__":
    main()